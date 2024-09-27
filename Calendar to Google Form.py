from __future__ import print_function
#Google API import
from apiclient import discovery
from httplib2 import Http
from oauth2client import client, file, tools

#Used to read ics, communicate with google calendar, and deal with logic when it comes to dates
from ics import Calendar, Event
from datetime import date, timedelta, datetime
import requests
import arrow

#Discord integration
from discordwebhook import Discord

#Importing excel functionality
import openpyxl

#GUI
import tkinter as tk
from tkinter import ttk
from tkinter import *

from dotenv import dotenv_values
config = dotenv_values(".env")

#Google Calendar URL
url = config["CALENDAR_URL"]

cal = Calendar(requests.get(url).text)
#print(cal.events)

spreadsheetPath = config["SPREADSHEET_PATH"]

#Making the GUI to display to users
allEntries = []
root = Tk()
frm = ttk.Frame(root, padding=20)
frm.grid()
def submit_response():
    global minDate 
    minDate = arrow.get(datetime(int(startYear.get()), int(startMonth.get()), int(startDay.get())))
    global maxDate
    maxDate = arrow.get(datetime(int(endYear.get()), int(endMonth.get()), int(endDay.get())))
    root.destroy()
def focus_next_widget_sMonth(event):
    if (len(startMonth.get())>1):
        event.widget.tk_focusNext().focus()
        return("break")
def focus_next_widget_sDay(event):
    if (len(startDay.get())>1):
        event.widget.tk_focusNext().focus()
        return("break")
def focus_next_widget_sYear(event):
    if (len(startYear.get())>3):
        event.widget.tk_focusNext().focus()
        return("break")
def focus_next_widget_eMonth(event):
    if (len(endMonth.get())>1):
        event.widget.tk_focusNext().focus()
        return("break")
def focus_next_widget_eDay(event):
    if (len(endDay.get())>1):
        event.widget.tk_focusNext().focus()
        return("break")
def focus_next_widget_eYear(event):
    if (len(endYear.get())>3):
        event.widget.tk_focusNext().focus()
        return("break")

Label(frm, text="Start Date (format: MM/DD/YYYY):").grid(column=0, row=0)
startMonth = Entry(frm, width=2)
startMonth.grid(column=1, row=0)
Label(frm, text="/").grid(column=2, row=0)
startDay = Entry(frm, width=2)
startDay.grid(column=3, row=0)
Label(frm, text="/").grid(column=4, row=0)
startYear = Entry(frm, width=4)
startYear.grid(column=5, row=0)
Label(frm, text="End Date (format: MM/DD/YYYY):").grid(column=0, row=1)
endMonth = Entry(frm, width=2)
endMonth.grid(column=1, row=1)
Label(frm, text="/").grid(column=2, row=1)
endDay = Entry(frm, width=2)
endDay.grid(column=3, row=1)
Label(frm, text="/").grid(column=4, row=1)
endYear = Entry(frm, width=4)
endYear.grid(column=5, row=1)
startMonth.focus()
for i in range(10):
    startMonth.bind("<KeyRelease-" + str(i) + ">", focus_next_widget_sMonth)
    startDay.bind("<KeyRelease-" + str(i) + ">", focus_next_widget_sDay)
    startYear.bind("<KeyRelease-" + str(i) + ">", focus_next_widget_sYear)
    endMonth.bind("<KeyRelease-" + str(i) + ">", focus_next_widget_eMonth)
    endDay.bind("<KeyRelease-" + str(i) + ">", focus_next_widget_eDay)
    endYear.bind("<KeyRelease-" + str(i) + ">", focus_next_widget_eYear)


submitButton = Button(frm, text="Submit", command=submit_response)
submitButton.grid(column=0, row=3)

root.mainloop()

rookieGameOptions = []
intIGameOptions = []

for event in cal.events:
    utc = arrow.get(event.begin)
    local = utc.to('US/Eastern')
    gDate = local.format("MM-DD-YYYY")
    month = int(local.format('MM'))
    day = int(local.format('DD'))
    year = int(local.format('YYYY'))
    print(arrow.get(datetime(year, month, day)).is_between(minDate,maxDate, '[)'))
    if (event.description.find("Rookies") != -1 or (event.description.find("Int. I") != -1 and event.description.find("Int. II") == -1)) and bool(arrow.get(datetime(year, month, day)).is_between(minDate,maxDate, '[)')):
        #print(local.format('YYYY-MM-DD HH:mm'))
        if int(local.format('HH')) < 12:
            time = str(local.format('HH:mm')) + "AM"
        elif int(local.format('HH')) == 12:
            time = str(local.format('HH:mm')) + "PM"
        else:
            time = str(int(local.format('HH'))-12) + local.format(':mm') + "PM"
        
        if (event.description.find("Rookies") != -1):
            rookieGameOptions.append(str("Game on " + gDate + " at " + time + " at " + event.location))
        else:
            intIGameOptions.append(str("Game on " + gDate + " at " + time + " at " + event.location))

#Sort games
rookieGameOptions.sort(reverse=True)
intIGameOptions.sort(reverse=True)

#GUI to confirm before uploading
googleFormsCheckButtonStatus1 = 0
discordCheckButtonStatus1 = 0

root1 = Tk()
frm1 = Frame(root1)
frm1.pack()

googleFormsCheckButtonStatus = IntVar()
discordCheckButtonStatus = IntVar()
intIButtonStatus = IntVar()
rookiesButtonStatus = IntVar()

def submit_response_2():
    global googleFormsCheckButtonStatus1
    global discordCheckButtonStatus1
    global intIButtonStatus1
    global rookiesButtonStatus1
    intIButtonStatus1 = int(intIButtonStatus.get())
    rookiesButtonStatus1 = int(rookiesButtonStatus.get())
    googleFormsCheckButtonStatus1 = int(googleFormsCheckButtonStatus.get())
    discordCheckButtonStatus1 = int(discordCheckButtonStatus.get())
    root1.destroy()

mainLabel = Label(text="Here are all the events that were found in the date range entered:")
mainLabel.pack()
if len(rookieGameOptions) != 0:
    rGamesLabel = Label(text="Rookie games: ")
    rGamesLabel.pack()
for i in reversed(range(len(rookieGameOptions))):
    Label(text=rookieGameOptions[i]).pack()
if len(intIGameOptions) != 0:
    Label(text="Int I games:").pack()
for i in reversed(range(len(intIGameOptions))):
    Label(text=intIGameOptions[i]).pack()

googleFormsCheckButton = Checkbutton(text="Send to Google Forms",variable=googleFormsCheckButtonStatus)
googleFormsCheckButton.pack()
googleFormsCheckButton.invoke()
discordCheckButton = Checkbutton(text="Send to Discord", variable=discordCheckButtonStatus)
discordCheckButton.pack()
discordCheckButton.invoke()
rookiesButton = Checkbutton(text="Include all Rookie Games", variable=rookiesButtonStatus)
rookiesButton.pack()
rookiesButton.invoke()
intIButton = Checkbutton(text="Include all Int I. Games", variable=intIButtonStatus)
intIButton.pack()
intIButton.invoke()
submitButton2 = Button(text="Submit", command=submit_response_2)
submitButton2.pack()

root1.mainloop()

#Google forms access
print(googleFormsCheckButtonStatus1)
if googleFormsCheckButtonStatus1 == 1:

    SCOPES = "https://www.googleapis.com/auth/forms.body"
    DISCOVERY_DOC = "https://forms.googleapis.com/$discovery/rest?version=v1"

    store = file.Storage('token.json')
    creds = None
    if not creds or creds.invalid:
        flow = client.flow_from_clientsecrets(config["CLIENT_SECRET"], SCOPES)
        creds = tools.run_flow(flow, store)

    form_service = discovery.build('forms', 'v1', http=creds.authorize(
        Http()), discoveryServiceUrl=DISCOVERY_DOC, static_discovery=False)

    # Request body to add a video item to a Form
    update = {
        "requests": [{
            "deleteItem" : {
                "location" : {
                    "index" : 0
                }
            }
            },
            {
            "deleteItem" : {
                "location" : {
                    "index" : 0
                }
            }
            },
            {
            "deleteItem" : {
                "location" : {
                    "index" : 0
                }
            }
            },
            {
            "createItem": {
                "item": {
                    "title": "What is your name?",
                    "questionItem": {
                        "question" : {
                            "required" : True,
                            "choiceQuestion" : {
                                "type" : "DROP_DOWN",
                                "options" : [],
                            }
                        }
                    }
                },
                "location": {
                    "index": 0
                }
            }
            },
            {
            "createItem": {
                "item": {
                    "title": "Rookie Baseball Games",
                    "description": "Games are 2 hours long",
                    "questionItem": {
                        "question" : {
                            "choiceQuestion" : {
                                "type" : "CHECKBOX",
                                "options" : [

                                ]
                            }
                        }
                    }
                },
                "location": {
                    "index": 1
                }
            }
            },
            {
            "createItem": {
                "item": {
                    "title": "Intermediate I Softball Games",
                    "description": "Games are 1.5 hours long",
                    "questionItem": {
                        "question" : {
                            "choiceQuestion" : {
                                "type" : "CHECKBOX",
                                "options" : [

                                ]
                            }
                        }
                    }
                },
                "location": {
                    "index": 2
                }
            }
            }
    ]}
    #print(update)
    #Adds game options into google form
    
    if (rookiesButtonStatus1 == 1):
        for i in reversed(range(len(rookieGameOptions))):
            update["requests"][4]["createItem"]["item"]["questionItem"]["question"]["choiceQuestion"]["options"].append({"value" : rookieGameOptions[i]})
    
    if (intIButtonStatus1 == 1):
        for i in reversed(range(len(intIGameOptions))):
            update["requests"][5]["createItem"]["item"]["questionItem"]["question"]["choiceQuestion"]["options"].append({"value" : intIGameOptions[i]})
        

    #if no events are there, create a placeholder to prevent error, or if user doesn't want to upload that type of game
    if len(rookieGameOptions) == 0 or rookiesButtonStatus1 == 0:
        update["requests"][4]["createItem"]["item"]["questionItem"]["question"]["choiceQuestion"]["options"].append({"value" : "Placeholder"})

    if len(intIGameOptions) == 0 or intIButtonStatus1 == 0:
        update["requests"][5]["createItem"]["item"]["questionItem"]["question"]["choiceQuestion"]["options"].append({"value" : "Placeholder"})

    #Update name selection on google forms to be based on excel
    wb = openpyxl.load_workbook(spreadsheetPath)
    ws = wb["Sheet2"]
    excelNames = [ws.cell(row=i,column=1).value for i in range(2, ws.max_row+1)]
    for name in excelNames:
        update["requests"][3]["createItem"]["item"]["questionItem"]["question"]["choiceQuestion"]["options"].append({"value" : str(name)})

    #print(update)
    #Debug statement
    #print(update["requests"][2]["createItem"]["item"]["questionItem"]["question"]["choiceQuestion"]["options"])


    form_id = config["FORM_ID"]

    # Change the form
    question_setting = form_service.forms().batchUpdate(
        formId=form_id, body=update).execute()

    # Print the result 
    result = form_service.forms().get(formId=form_id).execute()
    #print(result)

#Update the discord that the games for the following week has been posted
if discordCheckButtonStatus1 == 1:
    discord = Discord(url=config["DISCORD_WEBHOOK_URL"])
    discord.post(content="@everyone The availabilty form has been posted. Please fill it out and I will assign games in 24 hours.\nhttps://docs.google.com/forms/d/" + form_id +"/edit?pli=1")