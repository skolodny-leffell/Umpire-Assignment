from __future__ import print_function

#Google API import
from apiclient import discovery
from httplib2 import Http
from oauth2client import client, file, tools

#Excel integration
import openpyxl

#Discord integration
from discordwebhook import Discord

#GUI
import tkinter as tk
from tkinter import ttk
from tkinter import *

from dotenv import dotenv_values
config = dotenv_values(".env")

#Change to name of excel workbook
workbookName = config["SPREADSHEET_PATH"]

#Taking data from Google Forms in Json format (dictionary)
SCOPES = "https://www.googleapis.com/auth/forms.responses.readonly", "https://www.googleapis.com/auth/forms.body.readonly"
DISCOVERY_DOC = "https://forms.googleapis.com/$discovery/rest?version=v1"

store = file.Storage('token.json')
creds = None
if not creds or creds.invalid:
    flow = client.flow_from_clientsecrets(config["CLIENT_SECRET"], SCOPES)
    creds = tools.run_flow(flow, store)
service = discovery.build('forms', 'v1', http=creds.authorize(
    Http()), discoveryServiceUrl=DISCOVERY_DOC, static_discovery=False)

# Prints the responses of your specified form:
form_id = config["FORM_ID"]
responses = service.forms().responses().list(formId=form_id).execute()
#print(responses)
formBody = service.forms().get(formId=form_id).execute()
#print(formBody)



#print(responses["responses"])   
#print(formBody["items"][0]["questionItem"]["question"]["questionId"])
#print(formBody["items"][1])

#Getting only new responses from the Google Form and only taking data that is wanted

#Game submissions (nested tupple)
submissions = [[]]
#Names of umpires
names = []
#Count the amount of submissions (starts at 0)
submissionCounter = 0
#Id of question (unique)
questionId = []
#Game options
gameOptions = []

for i in formBody["items"]:
    questionId.append(i["questionItem"]["question"]["questionId"])
    #print(i["questionItem"]["question"]["questionId"])

for i in range(1, len(formBody["items"])):
    for j in formBody["items"][i]["questionItem"]["question"]["choiceQuestion"]["options"]:
        if j["value"].find("Placeholder") == -1:
            gameOptions.append(j["value"])

#Checks to see if name is first or answer is first
first = ""
for i in range(len(responses["responses"])):
    for j in responses["responses"][i]["answers"].keys():
        for k in responses["responses"][i]["answers"][j]["textAnswers"]["answers"]:
            #print(responses["responses"][i]["answers"][j]["questionId"])
            if responses["responses"][i]["answers"][j]["questionId"] == questionId[0] or responses["responses"][i]["answers"][j]["questionId"] == questionId[1] or responses["responses"][i]["answers"][j]["questionId"] == questionId[2]:
                if responses["responses"][i]["answers"][j]["questionId"] == questionId[0]:
                    if first == "":
                        first = "names"
                    #print("Name: " + k["value"])
                    names.append(k["value"])
                    #Doesn't go to next submission of games until there is a new umpire name
                    submissions.append([])
                    submissionCounter+=1
                elif k["value"].find("Placeholder") == -1 and (responses["responses"][i]["answers"][j]["questionId"] == questionId[1] or responses["responses"][i]["answers"][j]["questionId"] == questionId[2]):                
                    if first == "":
                        first = "answers"
                    #print("Submission " + k["value"])
                    submissions[submissionCounter].append(k["value"])

#If first value given is name, that means first index in submissions is blank, otherwise last index will be blank (and need to get delted)


if first == "names":
    submissions.pop(0)
else:
    submissions.pop(len(submissions)-1)
#print(names)
#print(submissions)


#Processing responses

#Name of assigned umpire (names are at same index at game they are assigned to in gameOptions)
assignedUmpireNames = []
secondAssignedUmpireNames = []

#List of all umpires
umpireList = []

#Makes tuple the same length as gameOptions - empty strings are placeholders
for i in gameOptions:
    assignedUmpireNames.append("")
    secondAssignedUmpireNames.append("")

#print(len(assignedUmpireNames))
#print(assignedUmpireNames)

def allGamesAssigned():
    for i in assignedUmpireNames:
        if i == "":
            return False
    return True

class Umpire:
    def __init__(self, name, availability):
        self.name = name
        self.availability = availability
        self.priority = 0
        self.trainee = False
        self.trainer = False


for i in range(len(names)):
    umpireList.append(Umpire(names[i], submissions[i]))

#Goes through Excel to find preset priority of umpires
wb = openpyxl.load_workbook(workbookName)
ws2 = wb["Sheet2"]
excelPriorities = [ws2.cell(row=i,column=7).value for i in range(2, ws2.max_row)]
excelNames = [ws2.cell(row=i,column=1).value for i in range(2, ws2.max_row)]

for umpire in umpireList:
    for i in range(len(excelNames)):
        if umpire.name == excelNames[i]:
            umpire.priority = excelPriorities[i]
            break
            
    


#Initial GUI (GUI 1) --> Gets inputs for umpire priority
root1 = Tk()
frm1 = Frame(root1)
frm1.grid()

umpirePriorityInputs = []
playoffGamesCheckButtonStatus = IntVar()
umpireTrainee = IntVar()
umpireTrainer = IntVar()
#Only used in GUI
umpireTraineeList = []
umpireTrainerList = []
for i in range(len(umpireList)):
    umpireTrainerList.append(IntVar(value=0))
    umpireTraineeList.append(IntVar(value=0))
#Final List - used outside GUI
umpireIsTraineeList = []
umpireIsTrainerList = []

def submit_response1():
    for i in range(len(umpireList)):
        umpireList[i].priority = int(umpirePriorityInputs[i].get())
        
        if int(umpireTraineeList[i].get()) == 1:
            umpireList[i].trainee = True
        if int(umpireTrainerList[i].get()) == 1:
            print("Trainer: " + umpireList[i].name)
            umpireList[i].trainer = True
    global playoffGamesCheckButtonStatus1
    playoffGamesCheckButtonStatus1 = int(playoffGamesCheckButtonStatus.get())   
    root1.destroy()

mainLabel = Label(frm1, text="Set the priority of all the umpires:")
mainLabel.grid(column=0,row=0)
Label(frm1, text="Umpire", font= ('Helvetica 15 underline')).grid(column=0,row=1)
Label(frm1, text="Priority", font= ('Helvetica 15 underline')).grid(column=1,row=1)
Label(frm1, text="Trainee", font= ('Helvetica 15 underline')).grid(column=2,row=1)
Label(frm1, text="Trainer", font= ('Helvetica 15 underline')).grid(column=3,row=1)

i = 2
for umpire in umpireList:
    umpireLabel = Label(frm1, text= str(umpire.name))
    umpireLabel.grid(column = 0, row = i)
    umpirePriorityInput = Entry(frm1, width=2)
    umpirePriorityInput.grid(column = 1, row = i)
    umpirePriorityInput.insert(END, umpire.priority)
    umpirePriorityInputs.append(umpirePriorityInput)
    umpireTraineeCheckButton = Checkbutton(frm1,variable=umpireTraineeList[i-2])
    umpireTraineeCheckButton.grid(column=2,row=i)
    umpireTrainerCheckButton = Checkbutton(frm1,variable=umpireTrainerList[i-2])
    umpireTrainerCheckButton.grid(column=3,row=i)
    i+=1

playoffGamesCheckButton = Checkbutton(frm1, text="Playoff game assignments (2 umpires, only based on priority)",variable=playoffGamesCheckButtonStatus)
playoffGamesCheckButton.grid(column=0, row=i)
submitButton1 = Button(frm1, text="Submit", command=submit_response1)
submitButton1.grid(column=0, row=i+1)

root1.mainloop()

#Sort the umpireList by priority

def takePriority(umpire):
    return umpire.priority

umpireList = sorted(umpireList, key=takePriority, reverse=True)

#Process of assigning games
#If this is a playoff game assigns 2 umpires just based on priority, otherwise assigns one umpire per game (partly based on priority, but focuses more on evening out how many games each person works)
if playoffGamesCheckButtonStatus1 == 1:
    for umpire in umpireList:
        for i in range(len(gameOptions)):
            for game in umpire.availability:
                if game == gameOptions[i]:
                    if assignedUmpireNames[i] == "":
                        assignedUmpireNames[i] = umpire.name
                    elif secondAssignedUmpireNames[i] == "":
                        secondAssignedUmpireNames[i] = umpire.name
                    break
                        
else:
    #Look at umpires who need training
    for i in range(len(umpireList)):
        if umpireList[i].trainee:
            print("Trainee")
            print(umpireList[i].name)
            alreadyAssigned = False
            for j in range(len(umpireList)):
                if alreadyAssigned:
                    break
                if umpireList[j].trainer:
                    print("Trainer")
                    print(umpireList[j])
                    if alreadyAssigned:
                        break
                    for k in range(len(gameOptions)):
                        for l in umpire.availability:
                            if alreadyAssigned:
                                break
                            if gameOptions[k] == l:
                                for m in umpire.availability:
                                    if gameOptions[k] == m:
                                        if assignedUmpireNames[k] == "" and secondAssignedUmpireNames[k] == "" and alreadyAssigned == False:
                                            assignedUmpireNames[k] = umpireList[j].name
                                            secondAssignedUmpireNames[k] = umpireList[i].name
                                            alreadyAssigned = True
                                            break
                        
                                        
                            



    #Look at people who only have 1 available game slot
    j = 0
    for umpire in umpireList:
        if len(umpire.availability) == 1 and umpire.trainee == False:
            for i in range(len(gameOptions)):
                if gameOptions[i] == umpire.availability[0]:
                    if assignedUmpireNames[i] == "":
                        assignedUmpireNames[i] = umpire.name
                        #print(assignedUmpireNames[i])
                    #If assigned or not assigned, go to next umpire (as already went through his singular available slot)
                    break
        j+=1

    #Look at games that only 1 person can umpire
    #Counts iterations of loop
    i = 0
    #Goes through each game
    for game in gameOptions:
        k = 0
        #keeps track of umpire that can do game name
        umpireAvailableName = ""
        #Counts how many people can umpire that game
        counter = 0
        #Goes through each umpire
        for umpire in umpireList:
            #Goes through all of the umpires availabilty and tests if equal to the game
            for j in range(len(umpire.availability)):
                if umpire.availability[j] == game and umpire.trainee == False:
                    counter+=1
                    #print(game)
                    #print(umpire.availability[j])
                    umpireAvailableName = umpire.name
                    #Go to next umpire as this umpire can already do the game in question
                    break
            #If multiple umpires can do the game, go to next game
            if counter >= 2:
                break
            k +=1
        #Runs only if one person can do the game and game is not already taken
        if counter == 1 and assignedUmpireNames[i] == "":
            assignedUmpireNames[i] = umpireAvailableName
        i+=1
                    

    #Look at people who have 0 games assigned in order of people with the least amount of available slots (that aren't booked) -->
    #to the people with the most amount of slots (that aren't booked)
    k = 0
    for umpire in umpireList:
        alreadyAssigned = False
        for assignedName in assignedUmpireNames:
            if assignedName == umpire.name:
                alreadyAssigned = True
                break
        if not alreadyAssigned and umpire.trainee == False:
            for i in range(len(gameOptions)):
                for j in umpire.availability:
                    if gameOptions[i] == j:
                        if assignedUmpireNames[i] == "":
                            assignedUmpireNames[i] = umpire.name
                            alreadyAssigned = True
                        break
                if alreadyAssigned:
                    break    
        k+=1 
            
    #print(assignedUmpireNames)
    #print(gameOptions)


    #Do the rest of the games based on seniority (priority) --> Give one umpire 2 and then next umpire 2, in order of senirority
    #If they have can't do any available games skip them, do this until all games assigned

    #for loop that breaks when assignedUmpireNames fills up (all games have been booked) or after 50 iterations
    
    for a in range(50):
        k = 0
        for umpire in umpireList:
            alreadyAssigned = False
            for i in range(len(gameOptions)):
                for j in umpire.availability:
                    if gameOptions[i] == j and umpire.trainee == False:
                        if assignedUmpireNames[i] == "":
                            assignedUmpireNames[i] = umpire.name
                            alreadyAssigned = True
                        break
                if alreadyAssigned:
                    break
            k+=1

        #If any games aren't assigned, goes back to beginning, if not, ends and goes to next part of code
        if allGamesAssigned():
                break

#Data sets needed for umpire chooser GUI
umpireGameAvailability = [] 
for i in gameOptions:
    umpireGameAvailability.append([])

for i in range(len(gameOptions)):
    for umpire in umpireList:
        for j in umpire.availability:
            if j == gameOptions[i]:
                umpireGameAvailability[i].append(umpire.name)



#Organize data that will be sent to Excel and Discord
gameDates = []
gameTimes = []
gameLocations = []

for game in gameOptions:
    gameDates.append(game[game.index("on")+3:game.index("on")+13])
    gameTimes.append(game[game.index("at")+3:game.index("M", game.index("at")+3)+1])
    gameLocations.append(game[game.index("at", game.index("at")+3)+3:])

#GUI #2 to change games that have already been assigned
root2 = Tk()
frm2 = Frame(root2)
frm2.grid()

umpireSelectors = []
secondUmpireSelectors = []
excelCheckButtonStatus = IntVar()
discordCheckButtonStatus = IntVar()

def submit_response2():
    for i in range(len(umpireSelectors)):
        assignedUmpireNames[i] = umpireSelectors[i].get()
        secondAssignedUmpireNames[i] = secondUmpireSelectors[i].get()
    global excelCheckButtonStatus1
    global discordCheckButtonStatus1
    excelCheckButtonStatus1 = int(excelCheckButtonStatus.get())
    discordCheckButtonStatus1 = int(discordCheckButtonStatus.get())
    root2.destroy()

#mainLabel = Label(frm2, text="Game Assignments:")
#mainLabel.grid(column=2,row=0)
Label(frm2, text="Umpire 1", font= ('Helvetica 15 underline')).grid(column=0,row=1)
Label(frm2, text="Umpire 2", font= ('Helvetica 15 underline')).grid(column=1,row=1)
Label(frm2, text="Game", font= ('Helvetica 15 underline')).grid(column=2,row=1)

finalRow = 3
for i in range(len(gameOptions)):
    clicked = StringVar()
    clicked1 = StringVar()
    clicked.set(assignedUmpireNames[i])
    clicked1.set(secondAssignedUmpireNames[i])
    umpireSelector = OptionMenu(frm2, clicked, *tuple(umpireGameAvailability[i]))
    umpireSelector.grid(column=0,row= i+2)
    umpireSelectors.append(clicked)
    secondUmpireSelector = OptionMenu(frm2, clicked1, *tuple(umpireGameAvailability[i]), "")
    secondUmpireSelector.grid(column=1, row = i+2)
    secondUmpireSelectors.append(clicked1)
    gameLabel = Label(frm2, text=gameOptions[i])
    gameLabel.grid(column=2,row=i+2)
    finalRow+=1

excelCheckButton = Checkbutton(frm2, text="Send to Excel",variable=excelCheckButtonStatus)
excelCheckButton.grid(row=finalRow+1,column=2)
excelCheckButton.invoke()
discordCheckButton = Checkbutton(frm2, text="Send to Discord", variable=discordCheckButtonStatus)
discordCheckButton.grid(row=finalRow+2,column=2)
discordCheckButton.invoke()
submitButton2 = Button(frm2, text="Submit", command=submit_response2)
submitButton2.grid(column=2, row=finalRow+3)
root2.mainloop()


#Export game assignments to Excel
if excelCheckButtonStatus1 == 1:
    ws1 = wb["Sheet1"]
    for i in range(len(assignedUmpireNames)):
        ws1.append([gameDates[i], gameTimes[i], gameLocations[i], assignedUmpireNames[i], secondAssignedUmpireNames[i]])
    wb.save(workbookName)



#Export game assignments to Discord
if discordCheckButtonStatus1 == 1:
    discordPost = "@everyone Here are the assigned games: "
    for i in range(len(assignedUmpireNames)):
        if secondAssignedUmpireNames[i] == "" or secondAssignedUmpireNames[i] == " ":
            discordPost = discordPost + "\n" + assignedUmpireNames[i] + "  " + gameDates[i] + "  " + gameTimes[i] + "  " + gameLocations[i]
        else:
            discordPost = discordPost + "\n" + assignedUmpireNames[i] + "  " + secondAssignedUmpireNames[i] + "  " + gameDates[i] + "  " + gameTimes[i] + "  " + gameLocations[i]
    discord = Discord(url=config["DISCORD_WEBHOOK_URL"])
    discord.post(content=discordPost)