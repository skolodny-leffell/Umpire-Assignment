import openpyxl
import os
import datetime
from docxtpl import DocxTemplate

from dotenv import dotenv_values
config = dotenv_values(".env")

#Classes
class Umpire:
    def __init__(self, name, gamesUmpired, payPerGame, totalPay, address):
        self.name = name
        self.gamesUmpired = gamesUmpired
        self.payPerGame = payPerGame
        self.totalPay = totalPay
        self.address = address

#Variables
#Stores umpire objects
umpireList = []

#name of excel workbook
workbookName = config["SPREADSHEET_PATH"]

#Dates
today = datetime.date.today()

year = today.year
month = today.month
day = today.day

#New Directory
parent_dir = "Umpire Assignment"
directory = "YBNR Check Request Forms " + str(year)

path = os.path.join(parent_dir, directory)

os.mkdir(path)

#Accessing excel
wb = openpyxl.load_workbook(workbookName, data_only=True)
ws = wb['Sheet2']
excelUmpireNames = [ws.cell(row=i, column=6).value for i in range(2, ws.max_row+1)]
gamesUmpired = [ws.cell(row=i, column=2).value for i in range(2, ws.max_row+1)]
payPerGame = [ws.cell(row=i, column=3).value for i in range(2, ws.max_row+1)]
totalPay = [ws.cell(row=i, column=4).value for i in range(2, ws.max_row+1)]
address = [ws.cell(row=i, column=5).value for i in range(2, ws.max_row+1)]

for i in range(len(excelUmpireNames)):
    umpire = Umpire(excelUmpireNames[i], gamesUmpired[i], payPerGame[i], totalPay[i], address[i])
    umpireList.append(umpire)

#Making Document
doc = DocxTemplate("Umpire Assignment\YBNR Check Request Form - Blank.docx")

for umpire in umpireList:
    if umpire.totalPay == 0 or umpire.totalPay == None:
        continue
    #print(umpire.address)
    if umpire.address.find(",") == -1:
        print("Address entered incorrectly for " + umpire.name)
        print("The current address in the system for him is " + umpire.address)
        address1 = umpire.address
        address2 = ""
    else:
        address1 = umpire.address[:umpire.address.index(",")]
        address2 = umpire.address[umpire.address.index(",")+2:]
    description = "Umpired " + str(umpire.gamesUmpired) + " Rookie/Int I. games at a rate of $" + str(umpire.payPerGame) + " per game"
    context = {'month' : month, 'day' : day, 'year' : year, 'money' : umpire.totalPay,
            'description' : description, 'name' : umpire.name, 'address1' : address1, 'address2' : address2}
    doc.render(context)
    doc.save("Umpire Assignment\YBNR Check Request Forms 2023\YBNR Check Request Form - " + umpire.name + ".docx")

